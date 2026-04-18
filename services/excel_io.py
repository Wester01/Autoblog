from openpyxl import load_workbook
from typing import List
from utils.types import EstructuraSEO
from pathlib import Path
import shutil
import os
import tempfile
import time

class ExcelIO:
    """
    Encargado de leer keywords de excel
    Encargado de escribir en excel la estructura soe H1 + H2
    """

    def __init__(self, file_path: str):
        self.path = Path(file_path)

        if not self.path.exists():
            raise FileNotFoundError(f"No existe el archivo Excel: {self.path}")


    #Lectura de keywords
    def keyword_reader(self) -> List[str]:
        wb = load_workbook(self.path)
        ws = wb.active #hoja única

        keywords = []
        fila = 2

        while True:
            celda = ws[f"D{fila}"].value

            if celda is None or str(celda).strip() == "":
                break

            keywords.append(str(celda).strip())
            fila+=1

        wb.close()

        if not keywords:
            raise ValueError("No se encontraron KW en la columna D")
        return keywords

    def structure_writer(self, estructura: EstructuraSEO):
        try:
            #backup del archivo
            backup_path = self.path.with_suffix(".backup.xlsx")
            shutil.copy2(self.path, backup_path)

            #carga del archivo
            wb = load_workbook(self.path)
            ws = wb.active

            # H1 en b2
            ws["B2"] = estructura["h1"]

            fila = 3
            for titulo in estructura["h2"]:
                ws[f"B{fila}"] = titulo
                fila += 1

            #Guardado en archivo temporal
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                temp_path = Path(tmp.name)

            wb.save(temp_path)
            wb.close()
            del wb
            #Reemplazo atómico
            os.replace(temp_path, self.path)
            time.sleep(0.5)
        except Exception as e:
            raise RuntimeError(f"Error escribiendo el documento Excel: {e}")



